import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

os.chdir('C:/Users/86185/OneDrive/桌面/新建文件夹 (2)')

# 读取目标文件
files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
target_file = [f for f in files if f.startswith('学员') or '事项' in f or f.startswith('ѧ')][0]
df = pd.read_excel(target_file)

# 添加辅助列
df['事项结束日期'] = df['事项时间（开始时间-结束时间）'].apply(lambda x: str(x).split('-')[3] + '-' + str(x).split('-')[4] + '-' + str(x).split('-')[5].split(' ')[0] if pd.notna(x) and len(str(x).split('-')) >= 6 else '')
df['核算状态'] = df.apply(lambda row: '未完成' if pd.isna(row['完成时间']) else ('已完成' if row['完成时间'] <= row['事项结束日期'] else '超时完成'), axis=1)
df['状态核验'] = df.apply(lambda row: '一致' if row['完成状态'] == row['核算状态'] else '不一致', axis=1)

# 保存为新的Excel文件
output_path = 'C:/Users/86185/.openclaw/workspace/学员事项完成情况_核算版.xlsx'

# 使用openpyxl创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "学员事项完成情况"

# 定义样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
align_center = Alignment(horizontal='center', vertical='center')

# 不一致行的样式（用于后续条件格式）
inconsistent_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
inconsistent_font = Font(color='9C0006')

# 写入表头
headers = list(df.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center

# 写入数据
for row_idx, row_data in enumerate(df.values, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = align_center
        # 如果不一致，标红
        if col_idx == len(headers) and value == '不一致':
            cell.fill = inconsistent_fill
            cell.font = inconsistent_font

# 设置列宽
column_widths = {
    'A': 8,   # 姓名
    'B': 14,  # 账号
    'C': 40,  # 部门
    'D': 20,  # 岗位
    'E': 25,  # 事项名称
    'F': 35,  # 事项时间
    'G': 10,  # 完成状态
    'H': 18,  # 完成时间
    'I': 12,  # 事项结束日期
    'J': 10,  # 核算状态
    'K': 10,  # 状态核验
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 添加自动筛选
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df)+1}"

# 冻结首行
ws.freeze_panes = 'A2'

# 添加说明工作表
ws_help = wb.create_sheet("使用说明")
ws_help['A1'] = "学员事项完成情况表 - 使用说明"
ws_help['A1'].font = Font(bold=True, size=14)

instructions = [
    "",
    "【表格说明】",
    "本表在原始数据基础上增加了三列辅助列，用于核验完成状态的准确性：",
    "",
    "1. 事项结束日期 (I列): 从事项时间中提取的截止日期",
    "2. 核算状态 (J列): 根据完成时间与截止日期重新计算的状态",
    "   - 未完成: 完成时间为空",
    "   - 已完成: 完成时间 ≤ 截止日期",
    "   - 超时完成: 完成时间 > 截止日期",
    "3. 状态核验 (K列): 比较原状态与核算状态是否一致",
    "   - 一致: 状态判断正确",
    "   - 不一致: 状态判断有误（标红显示）",
    "",
    "【使用方法】",
    "1. 筛选: 点击表头行的下拉箭头可进行筛选",
    "2. 检查不一致: 在K列筛选'不一致'，可发现状态判断错误的记录",
    "3. 统计数据: 选中整列可在Excel底部看到计数",
    "",
    "【数据统计】",
    f"总记录数: {len(df)} 条",
    f"按时完成: {len(df[df['完成状态']=='按时完成'])} 条",
    f"已完成: {len(df[df['完成状态']=='已完成'])} 条",
    f"未完成: {len(df[df['完成状态']=='未完成'])} 条",
    f"状态不一致: {len(df[df['状态核验']=='不一致'])} 条",
]

for i, text in enumerate(instructions, 2):
    ws_help[f'A{i}'] = text

ws_help.column_dimensions['A'].width = 70

# 保存
wb.save(output_path)
print(f"核算版Excel文件已保存到: {output_path}")
print(f"总行数: {len(df) + 1} (含表头)")
print(f"状态不一致记录数: {len(df[df['状态核验']=='不一致'])} 条")
