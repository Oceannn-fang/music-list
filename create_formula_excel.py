import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

os.chdir('C:/Users/86185/OneDrive/桌面/新建文件夹 (2)')

# 读取目标文件（学员事项完成情况）
files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
target_file = [f for f in files if f.startswith('学员') or '事项' in f or f.startswith('ѧ')][0]
df = pd.read_excel(target_file)

# 添加辅助列用于公式判断
df['事项结束时间'] = ''
df['公式判断状态'] = ''
df['是否一致'] = ''

# 保存为新的Excel文件
output_path = 'C:/Users/86185/.openclaw/workspace/学员事项完成情况_带公式.xlsx'

# 使用openpyxl创建带公式的工作簿
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "学员事项完成情况"

# 写入表头
headers = list(df.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 写入数据
for row_idx, row in enumerate(df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        if col_idx <= 8:  # 前8列是原始数据
            ws.cell(row=row_idx, column=col_idx, value=value)

# 添加公式列
# 假设：事项时间在F列，完成时间在H列
# I列提取结束时间（公式）
# J列用公式判断状态
# K列判断是否一致

for row_idx in range(2, len(df) + 2):
    # I列: 提取事项结束时间（使用MID和FIND函数）
    # 时间格式: 2024-12-23 08:00-2025-01-23 18:00
    # 提取后面部分: 2025-01-23 18:00
    ws.cell(row=row_idx, column=9).value = f'=MID(F{row_idx},FIND("-",F{row_idx},12)+1,LEN(F{row_idx}))'
    
    # J列: 公式判断状态
    # IF(H2="","未完成",IF(H2<=I2,"已完成","超时完成"))
    # 注意: 需要确保H列和I列是日期格式
    formula = f'=IF(H{row_idx}="","未完成",IF(H{row_idx}<=DATEVALUE(I{row_idx}),"已完成","超时完成"))'
    ws.cell(row=row_idx, column=10).value = formula
    
    # K列: 判断是否一致
    ws.cell(row=row_idx, column=11).value = f'=IF(G{row_idx}=J{row_idx},"一致","不一致")'

# 设置列宽
column_widths = [8, 12, 35, 20, 20, 35, 10, 18, 18, 10, 8]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 添加说明工作表
ws_help = wb.create_sheet("使用说明")
ws_help['A1'] = "完成状态判断公式说明"
ws_help['A1'].font = Font(bold=True, size=14)

ws_help['A3'] = "判断逻辑:"
ws_help['A4'] = "1. 如果完成时间为空 → 未完成"
ws_help['A5'] = "2. 如果完成时间 ≤ 事项结束时间 → 已完成"
ws_help['A6'] = "3. 如果完成时间 > 事项结束时间 → 超时完成"

ws_help['A8'] = "各列说明:"
ws_help['A9'] = "A-G列: 原始数据"
ws_help['A10'] = "I列: 通过公式从事项时间提取结束时间"
ws_help['A11'] = "J列: 通过公式根据完成时间和结束时间判断状态"
ws_help['A12'] = "K列: 比较原状态(G列)与公式判断状态(J列)是否一致"

ws_help['A14'] = "筛选方法:"
ws_help['A15'] = "1. 选中表头行"
ws_help['A16'] = "2. 数据 → 筛选"
ws_help['A17'] = "3. 可以在K列筛选'不一致'来检查异常数据"

# 设置说明列宽
ws_help.column_dimensions['A'].width = 60

# 保存
wb.save(output_path)
print(f"带公式的Excel文件已保存到: {output_path}")
print(f"总行数: {len(df) + 1} (含表头)")
